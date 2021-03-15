from PyQt5.QtWidgets import QMainWindow, QWidget, QApplication, QLabel, QPushButton, QAction, QComboBox, QCheckBox, QTableWidget, QHeaderView, QTableWidgetItem, QDialog, QMessageBox, QDialogButtonBox, QVBoxLayout, QHBoxLayout
from PyQt5 import QtCore
from PyQt5.QtGui import QIcon, QPixmap, QImage 
import PyQt5
import threading
import sys
import cv2
import datetime
import time
from openpyxl import load_workbook
import warnings

warnings.filterwarnings("ignore", category=FutureWarning)
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.title = "QRIT Reader"
        self.resize(500,700)
        self.move(300,30)
        self.setWindowTitle('QRIT Reader')
        self.setWindowIcon(QIcon('qrit100.png'))
        self.setFixedSize(500,600)
        self.setStyleSheet("background-color: white; color : red;")
        self.labelPowered = QLabel("Powered By : ", self)
        self.labelPowered.setGeometry(330, 20, 70, 40)

        self.logo = QLabel(self)
        self.logo.setGeometry(430,0, 70, 70)
        self.logopic = QPixmap('qrit100.png')
        self.logopic = self.logopic.scaled(70,70, QtCore.Qt.KeepAspectRatio, QtCore.Qt.FastTransformation)
        self.logo.setPixmap(self.logopic)

        global cameraviewer
        self.cameraviewer = QLabel(self)
        self.cameraviewer.setGeometry(QtCore.QRect(5,70, 490,350))
        self.cameraviewer.setStyleSheet("border : 2px solid black; border-top: 10px solid black; border-bottom: 10px solid black; ")
        global scanningLabel
        self.scanningLabel = QLabel("Loading...",self)
        self.scanningLabel.setGeometry(5, 420, 300, 20)
        self.scanningLabel.setStyleSheet("font-size : 15px")
        global detectedDataLabel
        self.detectedDataLabel = QLabel(self)
        self.detectedDataLabel.setGeometry(5, 442, 300, 120)
        self.detectedDataLabel.setStyleSheet("font-size : 15px")

        self.showFullRecordsButton = QPushButton("Show Full Records", self)
        self.showFullRecordsButton.setGeometry(400, 450, 100, 30)
        
        self.exitButton = QPushButton("Exit", self)
        self.exitButton.setGeometry(400, 500, 100, 30)
        self.exitButton.clicked.connect(exitThisApp)
        self.showFullRecordsButton.clicked.connect(switchToData)
        QtCore.QMetaObject.connectSlotsByName(self)
        
    def start_scanning(self, cameraIndex):
                
        camera = cv2.VideoCapture(cameraIndex)
        wb = load_workbook(filename = 'qritdata.xlsx')
        global runThread
        runThread = True
        
        qrdetector = cv2.QRCodeDetector()
        sheet = wb.active
        counter = sheet.max_row
        while  runThread:
            self.scanningLabel.setText("Scanning...")
            return_value , image  = camera.read()
            if return_value:
                image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
                img = QImage(image, image.shape[1], image.shape[0], QImage.Format_RGB888)
                pix = QPixmap.fromImage(img)
                pix = pix.scaled(490, 350, QtCore.Qt.KeepAspectRatioByExpanding, QtCore.Qt.SmoothTransformation)
                self.cameraviewer.setPixmap(pix)
            try:    
                decodedText, _ , _ = qrdetector.detectAndDecode(image)
            except:
                a = 0 
                
            if decodedText != "":
                
                counter = counter + 1
                data = decodedText.split("###")
                if len(data)==5:
                    self.scanningLabel.setText("Detected QR Code!! Saving Data!!")
                    self.detectedDataLabel.setText("Last Scanned\nFirst Name : "+ data[0] + "\nLast Name : "+data[1] + "\nCell No. : " + data[2] + "\nEmail : "+data[3]+ "\nCountry : "+data[4])
                    sheet["A"+str(counter)] = data[0] #First Name
                    sheet["B"+str(counter)] = data[1] #Last Name
                    sheet["C"+str(counter)] = data[2] #Cell Number
                    sheet["D"+str(counter)] = data[3] #Email
                    sheet["E"+str(counter)] = data[4] #Country
                    sheet["F"+str(counter)] = datetime.datetime.now().strftime("%x") #current date and time
                    sheet["G"+str(counter)] = datetime.datetime.now().strftime("%X") #current date and time
                    wb.save("qritdata.xlsx")
    QR Code is not generated with QRIT!!")
                time.sleep(5)
        camera.release()
    def closeEvent(self, *args, **kwargs):
        exitThisApp()  

class DataWindow(QMainWindow):    
    def __init__(self):
        super().__init__()
        self.resize(900, 600)
        self.move(100,100)
        self.setWindowTitle("QRIT - Check-in Records")
        self.setWindowIcon(QIcon('qrit100.png'))
        self.setStyleSheet("background-color: white; color : red;")
        self.setFixedSize(900,600)
        # self.menubar = self.menuBar()
        # self.exitAct = QAction('&Exit', self)
        # self.exitAct.setShortcut('Alt+F4')
        # self.exitAct.setStatusTip('Exit application')
         
        # fileMenu = self.menubar.addMenu('&File')
        # fileMenu.addAction(self.exitAct)
        # sortMenu = self.menubar.addMenu('&Sort')

        # self.sortingAction = QAction('&Advanced Sorting', self)
        # self.sortingAction.setShortcut('Ctrl+s')
        # self.sortingAction.setStatusTip('Sorting Settings')

        # sortMenu.addAction(self.sortingAction)
        # self.menubar.show()
        # self.central_widget = QWidget()
        # self.setCentralWidget(self.central_widget)
        # self.mainLayout = QGridLayout(self.central_widget)

        self.exitButton = QPushButton("Exit", self)
        self.exitButton.setGeometry(5,1,100,20)
        self.sortButton = QPushButton("Sort", self)
        self.sortButton.setGeometry(120,1,100,20)

        self.exitButton.clicked.connect(switchToMain)
        self.sortButton.clicked.connect(showSortingWindow)

        QtCore.QMetaObject.connectSlotsByName(self)
        self.contentWidget = QWidget(self)
        self.contentWidget.setGeometry(5, 25, 890, 570)
               

    def closeEvent(self, *args, **kwargs):
        self.close()
        window.show()     
class SortingWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.resize(300, 350)
        self.move(200,150)
        self.setWindowTitle("QRIT - Sorting Settings")
        self.setWindowIcon(QIcon('qrit100.png'))
        self.setStyleSheet("background-color: white; color : red;")
        self.setFixedSize(300,350)

        self.nameLabel = QLabel("Name : ", self)
        self.nameLabel.setGeometry(5,5, 55,20)
        self.nameCombo = QComboBox(self)
        self.nameCombo.setGeometry(70, 5, 220, 20)
        self.nameCombo.addItems(["Default","Ascending", "Descending"])

        self.countryLabel = QLabel("Countries : ", self)
        self.countryLabel.setGeometry(5, 30, 55, 20)
        self.countryCombo = QComboBox(self)
        self.countryCombo.setGeometry(70, 30,220, 20)
        
        self.startDateLabel = QLabel("Start Date : ", self)
        self.startDateLabel.setGeometry(5, 55, 55, 20)
        self.startYearCombo = QComboBox(self)
        self.startYearCombo.setGeometry(70, 55, 50, 20)
        self.startYearCombo.addItems(["{:02d}".format(x) for x in range(2019,2030)])
        self.startYearCombo.activated.connect(self.changedStartMonth)
        self.startMonthCombo = QComboBox(self)
        self.startMonthCombo.setGeometry(125, 55, 50, 20)
        self.startMonthCombo.addItems(["{:02d}".format(x) for x in range(1,13)])
        self.startMonthCombo.activated.connect(self.changedStartMonth)
        self.startDayCombo = QComboBox(self)
        self.startDayCombo.setGeometry(180, 55, 50, 20)
        self.startDayCombo.addItems(["{:02d}".format(x) for x in range(1,32)])

        self.endDateLabel = QLabel("End Date : ", self)
        self.endDateLabel.setGeometry(5, 80, 55, 20)
        self.endYearCombo = QComboBox(self)
        self.endYearCombo.setGeometry(70, 80, 50, 20)
        self.endYearCombo.addItems(["{:02d}".format(x) for x in range(2019,2030)])
        self.endYearCombo.activated.connect(self.changedEndMonth)
        self.endMonthCombo = QComboBox(self)
        self.endMonthCombo.setGeometry(125, 80, 50, 20)
        self.endMonthCombo.addItems(["{:02d}".format(x) for x in range(1,13)])
        self.endMonthCombo.activated.connect(self.changedEndMonth)
        self.endDayCombo = QComboBox(self)
        self.endDayCombo.setGeometry(180, 80, 50, 20)
        self.endDayCombo.addItems(["{:02d}".format(x) for x in range(1,32)])

        self.ignoreDate = QCheckBox("Ignore Date", self)
        self.ignoreDate.setGeometry(70, 110, 100, 20)
        self.sortButton = QPushButton("Sort", self)
        self.sortButton.setGeometry(200, 310, 40, 20)
        self.sortButton.clicked.connect(switchToSortedDataWidnow)
        self.cancelButton = QPushButton("Cancel", self)
        self.cancelButton.setGeometry(250, 310, 40, 20)
        self.cancelButton.clicked.connect(self.close)
        QtCore.QMetaObject.connectSlotsByName(self)
    def changedStartMonth(self):
        if str(self.startMonthCombo.currentText()) in ["01", "03", "05", "07", "08", "10", "12"]:
            self.startDayCombo.clear()
            self.startDayCombo.addItems(["{:02d}".format(x) for x in range(1,32)])
        elif str(self.startMonthCombo.currentText()) in ["04", "06","09", "11"] :
            self.startDayCombo.clear()
            self.startDayCombo.addItems(["{:02d}".format(x) for x in range(1,31)])
        else:
            if int(str(self.startYearCombo.currentText()))%4 == 0:
                self.startDayCombo.clear()
                self.startDayCombo.addItems(["{:02d}".format(x) for x in range(1,30)])
            else :
                self.startDayCombo.clear()
                self.startDayCombo.addItems(["{:02d}".format(x) for x in range(1,29)])
    def changedEndMonth(self):
        if str(self.endMonthCombo.currentText()) in ["01", "03", "05", "07", "08", "10", "12"]:
            self.endDayCombo.clear()
            self.endDayCombo.addItems(["{:02d}".format(x) for x in range(1,32)])
        elif str(self.endMonthCombo.currentText()) in ["04", "06","09", "11"] :
            self.endDayCombo.clear()
            self.endDayCombo.addItems(["{:02d}".format(x) for x in range(1,31)])
        else:
            if int(str(self.endYearCombo.currentText()))%4 == 0:
                self.endDayCombo.clear()
                self.endDayCombo.addItems(["{:02d}".format(x) for x in range(1,30)])
            else :
                self.endDayCombo.clear()
                self.endDayCombo.addItems(["{:02d}".format(x) for x in range(1,29)])

            

def loadData():
    global countries
    countries = []
    countries.append("Default")
    wb = load_workbook(filename = 'qritdata.xlsx')
    sheet = wb.active
    rows = sheet.max_row
    column = sheet.max_column
    data = QTableWidget()
    data.setRowCount(rows-1)
    data.setColumnCount(column)
    data.setHorizontalHeaderLabels(['First Name', 'Last Name', 'Cell Number', 'Email', 'Country', 'Date', 'Time'])
    data.horizontalHeader().setStyleSheet("color : red")
    data.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
    for i in range(0,rows):
        if i==0:
            continue
        for j in range(0, column):
            itemValue = sheet.cell(row=i+1, column=j+1)
            if j==4 and itemValue.value not in countries:
                countries.append(itemValue.value)
            data.setItem(i-1, j, QTableWidgetItem(str(itemValue.value)))
           
    data.resizeColumnsToContents()
    return data
def loadSortedData():
    wb = load_workbook(filename = 'qritdata.xlsx')
    sheet = wb.active
    rows = sheet.max_row
    column = sheet.max_column
    data = QTableWidget()
    data.setRowCount(rows)
    data.setColumnCount(column)
    data.setHorizontalHeaderLabels(['First Name', 'Last Name', 'Cell Number', 'Email', 'Country', 'Date', 'Time'])
    data.horizontalHeader().setStyleSheet("color : red")
    data.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
    startDate = datetime.datetime.strptime(""+sortingWindow.startMonthCombo.currentText()+"/"+sortingWindow.startDayCombo.currentText()+"/"+sortingWindow.startYearCombo.currentText(), "%m/%d/%Y")
    endDate = datetime.datetime.strptime(""+sortingWindow.endMonthCombo.currentText()+"/"+sortingWindow.endDayCombo.currentText()+"/"+sortingWindow.endYearCombo.currentText(), "%m/%d/%Y")
    rowCounter = 0
    for row in sheet.values:
        if rowCounter == 0:
            rowCounter = rowCounter+1
            continue
        recordDate = datetime.datetime.strptime(row[5], "%m/%d/%y")
        
        if row[4] not in countries:
            countries.append(row[4])
        if sortingWindow.countryCombo.currentText() == "Default" and sortingWindow.ignoreDate.isChecked() == True :
            data.setItem(rowCounter - 1, 0, QTableWidgetItem(row[0]))
            data.setItem(rowCounter - 1, 1, QTableWidgetItem(row[1]))
            data.setItem(rowCounter - 1, 2, QTableWidgetItem(row[2]))
            data.setItem(rowCounter - 1, 3, QTableWidgetItem(row[3]))
            data.setItem(rowCounter - 1, 4, QTableWidgetItem(row[4]))
            data.setItem(rowCounter - 1, 5, QTableWidgetItem(row[5]))
            data.setItem(rowCounter - 1, 6, QTableWidgetItem(row[6]))
            rowCounter = rowCounter + 1
        elif sortingWindow.countryCombo.currentText() == "Default" and sortingWindow.ignoreDate.isChecked() == False :
            if recordDate >= startDate and recordDate <= endDate:
                data.setItem(rowCounter - 1, 0, QTableWidgetItem(row[0]))
                data.setItem(rowCounter - 1, 1, QTableWidgetItem(row[1]))
                data.setItem(rowCounter - 1, 2, QTableWidgetItem(row[2]))
                data.setItem(rowCounter - 1, 3, QTableWidgetItem(row[3]))
                data.setItem(rowCounter - 1, 4, QTableWidgetItem(row[4]))
                data.setItem(rowCounter - 1, 5, QTableWidgetItem(row[5]))
                data.setItem(rowCounter - 1, 6, QTableWidgetItem(row[6]))
                rowCounter = rowCounter + 1
        elif not sortingWindow.countryCombo.currentText() == "Default" and sortingWindow.ignoreDate.isChecked() == True :
            if row[4] == sortingWindow.countryCombo.currentText():     
                data.setItem(rowCounter - 1, 0, QTableWidgetItem(row[0]))
                data.setItem(rowCounter - 1, 1, QTableWidgetItem(row[1]))
                data.setItem(rowCounter - 1, 2, QTableWidgetItem(row[2]))
                data.setItem(rowCounter - 1, 3, QTableWidgetItem(row[3]))
                data.setItem(rowCounter - 1, 4, QTableWidgetItem(row[4]))
                data.setItem(rowCounter - 1, 5, QTableWidgetItem(row[5]))
                data.setItem(rowCounter - 1, 6, QTableWidgetItem(row[6]))
                rowCounter = rowCounter + 1
        else :
            if row[4] == sortingWindow.countryCombo.currentText() and recordDate >= startDate and recordDate <= endDate:     
                data.setItem(rowCounter - 1, 0, QTableWidgetItem(row[0]))
                data.setItem(rowCounter - 1, 1, QTableWidgetItem(row[1]))
                data.setItem(rowCounter - 1, 2, QTableWidgetItem(row[2]))
                data.setItem(rowCounter - 1, 3, QTableWidgetItem(row[3]))
                data.setItem(rowCounter - 1, 4, QTableWidgetItem(row[4]))
                data.setItem(rowCounter - 1, 5, QTableWidgetItem(row[5]))
                data.setItem(rowCounter - 1, 6, QTableWidgetItem(row[6]))
                rowCounter = rowCounter + 1
        
    if sortingWindow.nameCombo.currentText() == "Ascending":
        data.sortByColumn(0, QtCore.Qt.AscendingOrder)
    elif sortingWindow.nameCombo.currentText() == "Descending":
        data.sortByColumn(0, QtCore.Qt.DescendingOrder)
    data.setRowCount(rowCounter - 1)       
    data.resizeColumnsToContents()
    return data
def returnCameraIndex():
    # checks the first 10 indexes.
    index = 0
    arr = []
    i = 10
    while i > 0:
        cap = cv2.VideoCapture(index)
        if cap.read()[0]:
            name = "Port # "+str(i+1)
            arr.append(name)
            cap.release()
        index += 1
        i -= 1
    if len(arr) == 0:
        noCameras = QDialog()
        noCamera.setWindowTitle("No Camera Found - QRIT Reader")
        noCamera.setGeometry(200,150,200, 70)
        dLayout = QVBoxLayout()

        infoLabel = QLabel("No Camera Found!!\nPlease Setup a Camera or fix drivers!!")
        dLayout.addWidget(infoLabel)
        bBox = QDialogButtonBox(QDialogButtonBox.Ok)
        bBox.accepted.connect(noCameras.accept)
        dLayout.addWidget(bBox)
        noCameras.setLayout(dLayout)
        noCameras.accepted.connect(exitThisApp)
        noCameras.exec_()
    elif len(arr) > 1:
        
        selectCamera = QDialog()
        selectCamera.setWindowTitle("Select Camera - QRIT Reader")
        selectCamera.setGeometry(200,150,200, 70)
        dialogLayout = QVBoxLayout()

        availableCameras = QComboBox()
        availableCameras.addItems(arr)
        dialogLayout.addWidget(availableCameras)

        buttons = QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        buttonBox = QDialogButtonBox(buttons)
        buttonBox.accepted.connect(selectCamera.accept)
        buttonBox.rejected.connect(selectCamera.reject)

        dialogLayout.addWidget(buttonBox)
        selectCamera.setLayout(dialogLayout)
        selectCamera.accepted.connect(lambda : setIndex(availableCameras.currentIndex()))
        selectCamera.exec_()
        return x
    else :
        return 0
def setIndex(i):
    global x
    x = i
def showSortingWindow():
    sortingWindow.show()

def switchToSortedDataWidnow():
    # dataWindow.mainLayout.takeAt(0).widget().deleteLater()
    # dataWindow.mainLayout.addWidget(loadSortedData())
    # dataWindow.mainLayout.update()
    # dataWindow.setCentralWidget(loadSortedData())
    contentlayout = QVBoxLayout()
    contentlayout.addWidget(loadSortedData())
    dataWindow.contentWidget.setLayout(contentlayout)
    sortingWindow.close()
    dataWindow.show()
def switchToData():
    # dataWindow.mainLayout.takeAt(0).widget().deleteLater()
    # dataWindow.mainLayout.addWidget(loadData())
    # dataWindow.setCentralWidget(loadData())
    # dataWindow.mainLayout.update()
    contentlayout = QVBoxLayout()
    contentlayout.addWidget(loadData())
    dataWindow.contentWidget.setLayout(contentlayout)
    window.hide()
    dataWindow.show()
def switchToMain():
    dataWindow.close()
    window.show()

def exitThisApp():
    global runThread
    runThread = False
    
    app.exit()
app = QApplication(sys.argv)
window = MainWindow()

dataWindow = DataWindow()
sortingWindow = SortingWindow()

# dataWindow.exitAct.triggered.connect(switchToMain)
# dataWindow.sortingAction.triggered.connect(sortingWindow.show)
# dataWindow.mainLayout.addWidget(loadData())
# dataWindow.setCentralWidget(loadData())
# dataWindow.contentlayout.addWidget(loadData())
# dataWindow.contentWidget.setLayout(dataWindow.contentlayout)
loadData()
sortingWindow.countryCombo.addItems(countries)

th = threading.Thread(target=window.start_scanning, args=(returnCameraIndex(),))
th.start()
window.show()
sys.exit(app.exec_())